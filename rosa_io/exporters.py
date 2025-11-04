"""Simple data exporters for ROSA session logging."""

from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping, Sequence

__all__ = [
    "export_csv",
    "export_json",
    "export_excel_row",
    "ensure_parent",
]


def ensure_parent(path: Path) -> None:
    """Create parent directories for the given file path if missing."""
    path.parent.mkdir(parents=True, exist_ok=True)


def export_csv(path: str, row: Mapping[str, object]) -> None:
    """Append a row to CSV, writing header automatically on first call."""
    file_path = Path(path)
    ensure_parent(file_path)
    row_dict = dict(row)
    with file_path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=row_dict.keys())
        if f.tell() == 0:
            writer.writeheader()
        writer.writerow(row_dict)


def export_json(path: str, row: Mapping[str, object]) -> None:
    """Append a JSONL row with timestamp metadata."""
    file_path = Path(path)
    ensure_parent(file_path)
    payload = {"ts": datetime.now().isoformat(), **dict(row)}
    with file_path.open("a", encoding="utf-8") as f:
        json.dump(payload, f)
        f.write("\n")


def export_excel_row(path: str, row: Mapping[str, object], columns: Sequence[str], sheet_name: str = "ROSA") -> None:
    """Append a row to an Excel workbook (single sheet) using openpyxl."""
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise RuntimeError("openpyxl is required to export Excel summaries.") from exc

    file_path = Path(path)
    ensure_parent(file_path)
    workbook = None
    worksheet = None
    if file_path.exists():
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            existing_header = [
                worksheet.cell(row=1, column=idx + 1).value for idx in range(max(worksheet.max_column, len(columns)))
            ]
            header_matches = existing_header[: len(columns)] == list(columns) and worksheet.max_column == len(columns)
            if not header_matches:
                sheet_index = workbook.sheetnames.index(sheet_name)
                workbook.remove(worksheet)
                worksheet = workbook.create_sheet(sheet_name, index=sheet_index)
        else:
            worksheet = workbook.create_sheet(sheet_name)
    if workbook is None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    if worksheet.max_row <= 1 and worksheet.max_column <= 1 and worksheet.cell(1, 1).value is None:
        worksheet.delete_rows(1)
        worksheet.append(list(columns))
    elif worksheet.max_row == 0:
        worksheet.append(list(columns))
    elif worksheet.max_row >= 1:
        current_header = [worksheet.cell(row=1, column=idx + 1).value for idx in range(len(columns))]
        if current_header != list(columns):
            worksheet.delete_rows(1, worksheet.max_row)
            worksheet.append(list(columns))

    data = [row.get(col, "") for col in columns]
    worksheet.append(data)
    workbook.save(file_path)
